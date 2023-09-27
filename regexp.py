import openpyxl
import report                                                                    # has function which does errors report
import all_brand_names                                                                  # dictionary with all our brands
import time
import re

start = time.time()

FILE_NAME = "BMXAA.xlsx"                                                          # enter the file name and sheet names
TM_SHEET_NAME = 'TM'
MAIN_SHEET_NAME = 'Sheet1'

workbook = openpyxl.open(FILE_NAME)
main_sheet = workbook[MAIN_SHEET_NAME]
tm_sheet = workbook[TM_SHEET_NAME]

max_brands_array_count = 5

brand_column_index = 2                                          # check the index of the column with the brand and error
error_column_index = 8

coordinate = {'x': 2, 'y': 2}                                                  # enter coordinates to display the report

codes_tabel = all_brand_names.codes_tabel

tm_tabel = {

}


def create_tm_table():                                           # creates table with brands from the 'TM' table (above)
    for (sup_brand, brand) in tm_sheet.iter_rows(min_row=2):
        tm_tabel[sup_brand.value.upper()] = []

    for (sup_brand, brand) in tm_sheet.iter_rows(min_row=2):
        tm_tabel[sup_brand.value.upper()].append(brand.value)


def replace(dict):                              # the main function that replaces the wrong article with the correct one
    count = 0
    for (id, sup_brand, sup_art_num, brand, art_num, comment, tm_flag, error, name) in main_sheet.iter_rows(min_row=2):

        if tm_tabel[sup_brand.value.upper()][0]:    # if the supplier's brand is matched with our brand - save our brand
            brand.value = tm_tabel[sup_brand.value.upper()][0]
        tm_flag.value = len(tm_tabel[sup_brand.value.upper()])           # create value for cell with brand names amount
        if error.value == 'Не распознан артикул':
            art_num.value = sup_art_num.value
            for key, value in dict.items():
                if brand.value == key:        # if main dictionary include brand from our file it replaces prefix or end
                    for item in value:

                        # if re.search(rf"^{item}[^a-zA-Z]", sup_art_num.value):
                        if re.search(rf"^{item}", sup_art_num.value):
                            new_art = re.sub(rf"^{item}", '/', sup_art_num.value)                               # prefix
                            art_num.value = new_art
                            comment.value = f'cut {item}'
                            count += 1

                        # elif re.search(rf"[^a-zA-Z]{item}$", sup_art_num.value):
                        elif re.search(rf"{item}$", sup_art_num.value):
                            new_art = re.sub(rf"{item}$", '/', sup_art_num.value)                                  # end
                            art_num.value = new_art
                            comment.value = f'cut {item}'
                            count += 1

        if 1 < len(tm_tabel[sup_brand.value.upper()]) < max_brands_array_count:
            for index, item in enumerate(tm_tabel[sup_brand.value.upper()]):
                if index != 0:                              # create new tables for brands that have more than one brand
                    sheet_names = workbook.sheetnames
                    if f'TM{index}' not in sheet_names:
                        workbook.create_sheet(f'TM{index}', 0)
                        workbook.save(FILE_NAME)
                    sheet = workbook[f'TM{index}']
                    max1 = sheet.max_row

                    sheet.cell(row=max1 + 1, column=4).value = tm_tabel[sup_brand.value.upper()][index]
                    sheet.cell(row=max1 + 1, column=1).value = id.value
                    sheet.cell(row=max1 + 1, column=2).value = sup_brand.value
                    sheet.cell(row=max1 + 1, column=3).value = sup_art_num.value
                    sheet.cell(row=max1 + 1, column=5).value = art_num.value
                    sheet.cell(row=max1 + 1, column=6).value = comment.value
    print(count)                                                                              # print number of replaces


try:
    create_tm_table()
    replace(codes_tabel)
    report.report_by_brands(workbook, main_sheet, coordinate, brand_column_index, error_column_index)             # on top
except Exception as err:
    print(err)

finally:
    workbook.save(FILE_NAME)
    workbook.close()

    end = time.time() - start
    print(end)

