from openpyxl.utils.cell import get_column_letter
from collections import defaultdict


def report_by_brands(workbook, sheet, coordinate, brand_column_index, error_column_index):

    if 'reports' in workbook.sheetnames:
        del workbook['reports']
    result_sheet = workbook.create_sheet('reports')

    def count_duplicate_values(column_index):

        value_counts = defaultdict(int)
        tm_is_wrong_brands = defaultdict(int)
        art_is_wrong_brands = defaultdict(int)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            cell_value = row[column_index - 1]  # Column index is 1-based
            value_counts[cell_value] += 1

            if row[column_index - 1] == 'Не распознан артикул':
                cell_value = row[brand_column_index - 1]  # Column index is 1-based
                art_is_wrong_brands[cell_value] += 1

            if row[column_index - 1] == 'Не распознана Торговая марка':
                cell_value = row[brand_column_index - 1]  # Column index is 1-based
                tm_is_wrong_brands[cell_value] += 1

        workbook.close()
        return value_counts, art_is_wrong_brands, tm_is_wrong_brands

    errors_report = count_duplicate_values(error_column_index)

    count_item = 0
    for value, count in errors_report[0].items():
        save(result_sheet, workbook, coordinate['x']+count_item, coordinate['y'], value, count)
        count_item += 1

    count_item = 0
    for value, count in errors_report[1].items():
        if count >= 100:
            save(result_sheet, workbook, coordinate['x']+count_item, coordinate['y']+3, value, count)
            count_item += 1

    count_item = 0
    for value, count in errors_report[2].items():
        if count >= 100:
            save(result_sheet, workbook, coordinate['x']+count_item, coordinate['y'] + 6, value, count)
            count_item += 1


def save(result_sheet, workbook, x, y, value, count):

    result_sheet.column_dimensions[get_column_letter(y)].width = 26
    result_sheet.column_dimensions[get_column_letter(y+1)].width = 6

    result_sheet.cell(row=x, column=y).value = value
    result_sheet.cell(row=x, column=y+1).value = count
